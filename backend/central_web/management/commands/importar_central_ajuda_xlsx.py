from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from openpyxl import load_workbook

from central_web.models import Modulo, Treinamento


MAPA_MODULOS = {
    1: "Cadastros",
    2: "Estoque",
    3: "Vendas",
    4: "Financeiro",
    5: "Transportes",
}


def limpar_header(valor):
    return str(valor or "").strip().lower()


class Command(BaseCommand):
    help = "Importa módulos e treinamentos do XLSX da Central de Ajuda"

    def add_arguments(self, parser):
        parser.add_argument("arquivo", type=str)
        parser.add_argument("--usuario-id", type=int, default=None)
        parser.add_argument("--limpar", action="store_true")

    def handle(self, *args, **options):
        arquivo = options["arquivo"]
        usuario_id = options.get("usuario_id")
        limpar = options.get("limpar")

        if limpar:
            Treinamento.objects.all().delete()
            Modulo.objects.all().delete()
            self.stdout.write("Dados antigos removidos.")

        usuario_padrao = None
        if usuario_id:
            User = get_user_model()
            usuario_padrao = User.objects.filter(id=usuario_id).first()

        wb = load_workbook(arquivo)
        ws = wb.active

        headers = [limpar_header(cell.value) for cell in ws[1]]
        self.stdout.write(f"Headers encontrados: {headers}")

        criados_modulos = 0
        atualizados_modulos = 0
        criados_treinamentos = 0
        atualizados_treinamentos = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            antigo_id = data.get("id")
            modulo_codigo = data.get("módulo") or data.get("modulo")
            titulo = data.get("titulo")
            conteudo = data.get("conteudo")
            data_criacao = data.get("data de criação")
            data_atualizacao = data.get("data de atualização")
            video = data.get("video")

            if not antigo_id or not modulo_codigo or not titulo:
                continue

            modulo_codigo = int(modulo_codigo)
            modulo_nome = MAPA_MODULOS.get(modulo_codigo, f"Módulo {modulo_codigo}")

            modulo, modulo_created = Modulo.objects.update_or_create(
                id=modulo_codigo,
                defaults={
                    "nome": modulo_nome,
                    "descricao": "-",
                }
            )

            if modulo_created:
                criados_modulos += 1
            else:
                atualizados_modulos += 1

            treinamento, treinamento_created = Treinamento.objects.update_or_create(
                id=int(antigo_id),
                defaults={
                    "empresa": 1,
                    "modulo_id": modulo_codigo,
                    "titulo": str(titulo).strip(),
                    "conteudo": conteudo or "",
                    "data_criacao": data_criacao,
                    "data_atualizacao": data_atualizacao,
                    "usuario_criador": usuario_padrao,
                    "video": str(video).strip() if video else None,
                }
            )

            if treinamento_created:
                criados_treinamentos += 1
            else:
                atualizados_treinamentos += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Importação concluída. "
                f"Módulos criados: {criados_modulos} | "
                f"Módulos atualizados: {atualizados_modulos} | "
                f"Treinamentos criados: {criados_treinamentos} | "
                f"Treinamentos atualizados: {atualizados_treinamentos}"
            )
        )